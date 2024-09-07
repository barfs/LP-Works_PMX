# romero@engineer.com

# Package Requirements:
# pip install xlsxwriter pandas openpyxl xlwt

# How a message in a .msg file is formatted:
#  ID  M1      M2      M3      M4      M5      M6      M7      M8
# [__][___][_][___][_][___][_][___][_][___][_][___][_][___][_][___][_]
#  4   50   1  50   1  50   1  50   1  50   1  50   1  50   1  50   1

# Convert a .msg file to a spreadsheet file like so:
# - Take in the msg file given by the user
#   - If it exists, begin parsing 412 bytes:
#     - Check the first four bytes to get the message ID number (ID number X)
#       - Number is ABCD, where A is the least significant digit and D is the most
#       - ID = A + (B * 256) + (C * 256^2) + (D * 256^3) where A/B/C/D can range from 0 to 255
#     - Check the next fifty bits for the contents of message line 1 of message number X
#       - Any instance of a NULL byte indicates it is blank from that point on.
#     - Skip NULL byte 1
#     - Repeat for message lines and the individual NULL bytes 2 through 8
#     - Make the message out of the collected data
#     - Add message to list of messages
#       - If there are more bytes, repeat
#       - If there are no more bytes, proceed to export to an Excel spreadhseet
#         - Set header (first row)
#         - For each message
#           - Fill row with ID, M1 through M8, and merged messages
#           - Export as same-named file .xlsx
#         - Check if the IDs are in order
#           - If not in order, export a sorted version
#           - If in order, notify that it already is sorted
#   - If the file does not exist, notify and exit.

# The "problem" characters when using ord():
# 13: Carraige Return     ASCII Code: 10 (Interpretted as Line Feed)
# 128: €                  ASCII Code: 8364
# 129: � (Unused)         ASCII Code: 65533
# 130: ‚                  ASCII Code: 8218
# 131: ƒ                  ASCII Code: 402
# 132: „                  ASCII Code: 8222
# 133: …                  ASCII Code: 8230
# 134: †                  ASCII Code: 8224
# 135: ‡                  ASCII Code: 8225
# 136: ˆ                  ASCII Code: 710
# 137: ‰                  ASCII Code: 8240
# 138: Š                  ASCII Code: 352
# 139: ‹                  ASCII Code: 8249
# 140: Œ                  ASCII Code: 338
# 141: � (Unused)         ASCII Code: 65533
# 142: Ž                  ASCII Code: 381
# 143: � (Unused)         ASCII Code: 65533
# 144: � (Unused)         ASCII Code: 65533
# 145: ‘                  ASCII Code: 8216
# 146: ’                  ASCII Code: 8217
# 147: “                  ASCII Code: 8220
# 148: ”                  ASCII Code: 8221
# 149: •                  ASCII Code: 8226
# 150: –                  ASCII Code: 8211
# 151: —                  ASCII Code: 8212
# 152: ˜                  ASCII Code: 732
# 153: ™                  ASCII Code: 8482
# 154: š                  ASCII Code: 353
# 155: ›                  ASCII Code: 8250
# 156: œ                  ASCII Code: 339
# 157: � (Unused)         ASCII Code: 65533
# 158: ž                  ASCII Code: 382
# 159: Ÿ                  ASCII Code: 376

import xlsxwriter
import os
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import glob
import xlwt
import shlex
import tkinter as tk
from tkinter import filedialog, messagebox

encoding = "windows-1252"

message_chunk_bytes = 412
id_bytes = 4
message_line_bytes = 50

def empty_check(x):
    if len(x) > 1:
        return False
    elif len(x) == 1 and ord(x) == 0:
        return True
    else:
        return x == "" or len(x) == 0

def list_characters_with_ascii(input_string):
    lines = []
    for char in input_string:
        ascii_code = ord(char)
        line = f"{char}: {ascii_code}"
        lines.append(line)
    return "\n".join(lines)

class Message:
    def __init__(self, id=0, M1="", M2="", M3="", M4="", M5="", M6="", M7="", M8=""):
        self.id = id
        self.M1 = M1
        self.M2 = M2
        self.M3 = M3
        self.M4 = M4
        self.M5 = M5
        self.M6 = M6
        self.M7 = M7
        self.M8 = M8
        
    def set_id(self, id):
        self.id = id
        
    def set_m1(self, m1):
        self.M1 = m1
        
    def set_m2(self, m2):
        self.M2 = m2
        
    def set_m3(self, m3):
        self.M3 = m3
        
    def set_m4(self, m4):
        self.M4 = m4
        
    def set_m5(self, m5):
        self.M5 = m5
        
    def set_m6(self, m6):
        self.M6 = m6
        
    def set_m7(self, m7):
        self.M7 = m7
        
    def set_m8(self, m8):
        self.M8 = m8

    def valid_message_id(self):
        return self.id > 0
    
    def message_empty(self):
        return self.M1 != "" | self.M2 != "" | self.M3 != "" | self.M4 != "" | self.M5 != "" | self.M6 != "" | self.M7 != "" | self.M8 != ""
    
    def print(self):
        print("Message", self.id, ": M1: \"", self.M1,
              "\"  M2: \"", self.M2,
              "\"  M3: \"", self.M3,
              "\"  M4: \"", self.M4,
              "\"  M5: \"", self.M5,
              "\"  M6: \"", self.M6, 
              "\"  M7: \"", self.M7,
              "\"  M8: \"", self.M8)
    
    def merge_ms(self, separator=""):
        return f"{self.M1}{separator}{self.M2}{separator}{self.M3}{separator}{self.M4}{separator}{self.M5}{separator}{self.M6}{separator}{self.M7}{separator}{self.M8}"#.replace("\x00", "")

    def merge_ms_optimal(self, separator=""):

        merged_message_lines = self.merge_ms(separator)
        
        if separator == "" and len(separator) == 0:
            return merged_message_lines.replace("\x00", "")

        final_empty_flag = False
        
        # Remove any hanging blank lines
        if empty_check(self.M8):
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M8))]
        else:
            final_empty_flag = True
        if empty_check(self.M7) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M7))]
        else:
            final_empty_flag = True
        if empty_check(self.M6) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M6))]
        else:
            final_empty_flag = True
        if empty_check(self.M5) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M5))]
        else:
            final_empty_flag = True
        if empty_check(self.M4) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M4))]
        else:
            final_empty_flag = True 
        if empty_check(self.M3) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M3))]
        else:
            final_empty_flag = True 
        if empty_check(self.M2) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M2))]
        else:
            final_empty_flag = True 
        if empty_check(self.M1) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M1))]
        else:
            final_empty_flag = True
        
        return merged_message_lines.replace("\x00", "")
            
def get_id_from_bytes(A, B, C, D):
    return ord(A) + (256 * ord(B)) + (256 * 256 * ord(C)) + (256 * 256 * 256 * ord(D))

def get_id_from_ints(A, B, C, D):
    return A + (256 * B) + (256 * 256 * C) + (256 * 256 * 256 * D)

def write_to_file_m_format(file, message):
    if len(message) > message_line_bytes:
        message = message[:message_line_bytes]
    for i in range(message_line_bytes):
        if i < len(message):
            file.write(message[i].encode(encoding))
        else:
            file.write(b'\x00')

# NULL separator to previous message line number guide:
null_separator_pos_prev_m = {
    54: "M1",
    105: "M2",
    156: "M3",
    207: "M4",
    258: "M5",
    309: "M6",
    360: "M7",
    411: "M8"
}

def is_word_in_input(word, text):
    normalized_word = word.lower().replace(" ", "")
    normalized_text = text.lower().replace(" ", "")
    return normalized_word in normalized_text

print("PLU/MSG/Excel Python program started.\n")

debug_mode = False

# original_sf_warning_settings = warnings.simplefilter('default')
# original_fw_warning_settings = warnings.filterwarnings('default')
# original_warning_settings = warnings.filters[:]


# ------------------------------

def convert_msg_to_excel():

    msg_file_path = filedialog.askopenfilename(
        title="Select .msg File",
        filetypes=[("Message Files", "*.msg")]
    )
    
    if not msg_file_path:
        messagebox.showerror("Error", "Message file must be in the .msg format.")
        return
    
    msg_name, msg_extension = os.path.splitext(msg_file_path)
    msg_file_name_without_extension = os.path.basename(msg_name)

    save_file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")],
        title="Save Excel File"
    )
    
    if not save_file_path:
        messagebox.showerror("Error", "No save location selected!")
        return
     
    
    try:
        # Read .msg file and process it
        with open(msg_file_path, 'rb') as file:
            messages = []
            i = 0  # Byte counter
            A = B = C = D = 0  # ID
            message = Message()
            message_line = ""
            null_flag = False
            while True:
                byte = file.read(1)
                if not byte:
                    if message.valid_message_id():  # If there's a valid message when EOF is hit
                        messages.append(message)
                    break  # End of file
                integer_value = ord(byte)
                byte_char = chr(integer_value)
                
                # Position logic
                if i == message_chunk_bytes:  # Finished with message, reset
                    A = B = C = D = i = 0
                    messages.append(message)
                    message = Message()
                    message_line = ""
                    null_flag = False

                if i in null_separator_pos_prev_m:
                    if not null_flag and integer_value != 0:
                        message_line += byte_char
                    setattr(message, null_separator_pos_prev_m[i], message_line.replace("\x00", ""))
                    null_flag = False
                    message_line = ""
                    
                if not null_flag:
                    if i == 0:
                        A = integer_value
                    elif i == 1:
                        B = integer_value
                    elif i == 2:
                        C = integer_value
                    elif i == 3:
                        D = integer_value
                        message.set_id(get_id_from_ints(A, B, C, D))
                    elif i not in null_separator_pos_prev_m: 
                        if integer_value == 0:
                            null_flag = True
                        else:
                            message_line += byte_char
                i += 1

        def export_messages_to_excel(messages, file_path):
            workbook = xlsxwriter.Workbook(file_path)
            worksheet = workbook.add_worksheet()
            text_wrap_format = workbook.add_format({'text_wrap': True})
            header = ["ID", "M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "Merged", "Merged with Spaces", "Merged with Newlines"]
            for col, label in enumerate(header):
                worksheet.write(0, col, label)

            for row, message in enumerate(messages):
                worksheet.write(row + 1, 0, message.id)
                for col, value in enumerate([message.M1, message.M2, message.M3, message.M4, message.M5, message.M6, message.M7, message.M8, Message.merge_ms_optimal(message, ""), Message.merge_ms_optimal(message, " "), Message.merge_ms_optimal(message, "\n")]):
                    worksheet.write(row + 1, col + 1, value, text_wrap_format)
            workbook.close()

        # Export messages to Excel
        export_messages_to_excel(messages, save_file_path)

        is_sorted = all(messages[i].id <= messages[i + 1].id for i in range(len(messages) - 1))

        if is_sorted:
            messagebox.showinfo("Info", "The messages were already sorted by ID.")
        else:
            sorted_file_path = save_file_path.replace(".xlsx", " sorted.xlsx")
            sorted_messages = sorted(messages, key=lambda message: message.id)
            export_messages_to_excel(sorted_messages, sorted_file_path)
            messagebox.showinfo("Info", f"Messages were unsorted. A sorted version has been saved to: {sorted_file_path}")

        # Formatting notification
        messagebox.showinfo(
            "Formatting Tip", 
            "To properly display the merged-text results, select the columns and use the \"Wrap Text\" option in Excel.\n"
            "You may also need to adjust vertical alignment (top or middle) for better display."
        )

    except FileNotFoundError:
        messagebox.showerror("Error", f"{msg_file_path} not found. Ensure the file exists.")
        return
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")
        return

def convert_excel_to_msg():
    excel_file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    
    if not excel_file_path:
        messagebox.showerror("Error", "No Excel file selected!")
        return

    excel_name, excel_extension = os.path.splitext(excel_file_path)
    excel_file_name_without_extension = os.path.basename(excel_name)

    save_file_path = filedialog.asksaveasfilename(
        defaultextension=".msg",
        filetypes=[("Message Files", "*.msg")],
        title="Save .msg File"
    )
    
    if not save_file_path:
        messagebox.showerror("Error", "No save location selected!")
        return

    if os.path.exists(save_file_path):
        overwrite = messagebox.askyesno("Overwrite", f"The file '{save_file_path}' already exists. Do you want to overwrite it?")
        if not overwrite:
            return

    try:
        # Read Excel file
        df = pd.read_excel(excel_file_path)

        # Drop rows with missing ID (this would cause issues for the .msg file)
        df = df.dropna(subset=['ID'])
        rows = 0
        messages = []
        ids = []

        for index, row in df.iterrows():
            e_message_id = int(row['ID'])
            e_ms = [str(row[f'M{i}']) if pd.notna(row[f'M{i}']) else '' for i in range(1, 9)]
            if len(e_ms) > message_line_bytes:  # Message lines cannot exceed 50 characters
                e_ms = e_ms[:message_line_bytes]
            ids.append(e_message_id)
            messages.append(Message(e_message_id, *e_ms))
            if debug_mode:
                print(f"[{rows}] Message ID: {e_message_id}")
                print(f"[{rows}] Messages 1-8: {e_ms}")

        used_ids = []
        repeat_ids = []

        if len(messages) > 0:
            with open(save_file_path, 'wb') as file:
                for msg in messages:
                    A = B = C = D = 0  # ID
                    if debug_mode:
                        print(f"Message Info: [{msg.id}] \"{msg.M1}\" \"{msg.M2}\" \"{msg.M3}\" \"{msg.M4}\" \"{msg.M5}\" \"{msg.M6}\" \"{msg.M7}\" \"{msg.M8}\"")
                    if msg.id not in used_ids:
                        if msg.id > 4294967295:
                            msg.id = 4294967295
                        elif msg.id < 1:
                            msg.id = 1
                        D = msg.id // (256 ** 3)
                        C = (msg.id % (256 ** 3)) // (256 ** 2)
                        B = (msg.id % (256 ** 2)) // 256
                        A = msg.id % 256
                        if debug_mode:
                            print(f"Message ID Base-10 to Base-256: {D} | {C} | {B} | {A}\n")
                        #412: 4 50 1 50 1 50 1 50 1 50 1 50 1 50 1 50 1
                            file.write(bytes([A]))                     # 1    |     1
                            file.write(bytes([B]))                     # 1    |     2
                            file.write(bytes([C]))                     # 1    |     3
                            file.write(bytes([D]))                     # 1    |     4
                            write_to_file_m_format(file, msg.M1)       # 50   |    54
                            file.write(b'\x00')                        # 1    |    55
                            write_to_file_m_format(file, msg.M2)       # 50   |   105
                            file.write(b'\x00')                        # 1    |   106
                            write_to_file_m_format(file, msg.M3)       # 50   |   156
                            file.write(b'\x00')                        # 1    |   157
                            write_to_file_m_format(file, msg.M4)       # 50   |   207
                            file.write(b'\x00')                        # 1    |   208
                            write_to_file_m_format(file, msg.M5)       # 50   |   258
                            file.write(b'\x00')                        # 1    |   259
                            write_to_file_m_format(file, msg.M6)       # 50   |   309
                            file.write(b'\x00')                        # 1    |   310
                            write_to_file_m_format(file, msg.M7)       # 50   |   360
                            file.write(b'\x00')                        # 1    |   361
                            write_to_file_m_format(file, msg.M8)       # 50   |   411
                            file.write(b'\x00')                        # 1    |   412

                        used_ids.append(msg.id)
                    else:
                        repeat_ids.append(msg.id)

        if len(used_ids) < len(ids) or repeat_ids:
            messagebox.showwarning("Warning", f"Some messages had reused IDs and were not included: {repeat_ids}")

        # Completion message
        messagebox.showinfo("Success", f"File saved to {save_file_path}")

    except FileNotFoundError:
        messagebox.showerror("Error", f"{excel_file_path} not found. Ensure the file exists.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

    proceed = True
    if debug_mode:
        # Unsuppress FutureWarnings
        warnings.simplefilter('default')
    else:
        # Suppress FutureWarnings
        with warnings.catch_warnings():
            warnings.simplefilter(action='ignore', category=FutureWarning)
    
    print("NOTE: This merging action does NOT support merging formulas.")
    print("Merged-message cells using formulas will say \'0\' instead, but all message lines (M1 - M8) will remain intact.\n")

    plu_file_path = str(input("Enter the PLU (.xlsx) file to add the messages to: "))
    plu_name, plu_extension = os.path.splitext(plu_file_path)
    plu_file_name_without_extension = os.path.basename(plu_name)
    print("")
    msg_file_path = str(input("Enter the msg (.xlsx) file be added: "))
    msg_name, msg_extension = os.path.splitext(msg_file_path)
    msg_file_name_without_extension = os.path.basename(msg_name)
    print("")

    def merge_files(plu_file, msg_file):
        try:
            # Create a dictionary to map IDs to corresponding message data
            id_to_message = {row['ID']: row[1:] for _, row in msg_file.iterrows()}

            msg_headers = ["M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "Merged", "Merged with Spaces", "Merged with Newlines"]
            plu_file = plu_file.reindex(columns=[*plu_file.columns, *msg_headers])

            # Iterate through plu_file and update the new columns based on MessageNo
            for index, row in plu_file.iterrows():
                message_no = row['MessageNo']
                if message_no > 0 and message_no in id_to_message:
                    message_data = id_to_message[message_no]
                    for i, data in enumerate(message_data):
                        if str(data) != "nan":
                            # Convert the column to 'object' type
                            plu_file[msg_headers[i]] = plu_file[msg_headers[i]].astype(object)
                            # Update the DataFrame
                            plu_file.at[index, msg_headers[i]] = str(data)

            return plu_file

        except FileNotFoundError:
            print("Error:", plu_file_path, "or", msg_file_path, "not found. Make sure path is correct, file is in the correct directory, or file exists.")
            proceed = False

    if proceed:
        try:
            # Load the first Excel file (plu_file_path)
            plu_file = pd.read_excel(plu_file_path)

            # Load the second Excel file (msg_file_path)
            msg_file = pd.read_excel(msg_file_path)

        except FileNotFoundError:
            print("Error:", plu_file_path, "or", msg_file_path, "not found. Make sure path is correct, file is in the correct directory, or file exists.")
            proceed = False

    if proceed:
        merged_file = merge_files(plu_file, msg_file)

        new_file_path = plu_file_name_without_extension + '+' + msg_file_name_without_extension + '.xlsx'

        if os.path.exists(new_file_path):
            user_ow_input = input(f"The file \'{new_file_path}\' already exists. Do you want to overwrite it? (y/n): ").lower()
            if user_ow_input == 'y' or user_ow_input == "yes":
                try:
                    os.remove(new_file_path)
                    print(f"Overwriting {new_file_path} ...")
                except FileNotFoundError:
                    print(f"Error: File {new_file_path} not found.")
                except PermissionError:
                    print(f"Error: Permission denied to delete file {new_file_path}.")
                    proceed = False
                except Exception as e:
                    print(f"Error: An unknown error occurred. Exception: {e}")
                    proceed = False
            elif user_ow_input == 'n' or user_ow_input == "no":
                print("Overwriting canceled.")
                proceed = False
            else:
                print("Unknown input. Operation and overwriting canceled. The existing file was not overwritten.")
                proceed = False

    if proceed:    
        try:
            merged_file.to_excel(new_file_path, index=False)
        except PermissionError:
            print("Permission Error: Close", new_file_path, "if open.")
            proceed = False

    if proceed:
        workbook = openpyxl.load_workbook(new_file_path)
        worksheet = workbook['Sheet1']
        header_row = worksheet[1]
        header_to_column = {cell.value: cell.column for cell in header_row}
        columns_to_wrap = ["Merged", "Merged with Spaces", "Merged with Newlines"]
        for header_name in columns_to_wrap:
            column_index = header_to_column.get(header_name)
            if column_index:
                for cell in worksheet[column_index]:
                    cell.alignment = openpyxl.styles.Alignment(wrapText=True)
        workbook.save(new_file_path)
        print(f"Merged and saved to {new_file_path}.\n")

def append_msg_to_plu():
    plu_file_path = filedialog.askopenfilename(
        title="Select PLU Excel File",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    
    if not plu_file_path:
        messagebox.showerror("Error", "No PLU file selected!")
        return

    plu_name, plu_extension = os.path.splitext(plu_file_path)
    plu_file_name_without_extension = os.path.basename(plu_name)

    msg_file_path = filedialog.askopenfilename(
        title="Select Message Excel File",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    
    if not msg_file_path:
        messagebox.showerror("Error", "No Message file selected!")
        return

    msg_name, msg_extension = os.path.splitext(msg_file_path)
    msg_file_name_without_extension = os.path.basename(msg_name)

    def merge_files(plu_file, msg_file):
        try:
            # Create a dictionary to map IDs to corresponding message data
            id_to_message = {row['ID']: row[1:] for _, row in msg_file.iterrows()}

            msg_headers = ["M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "Merged", "Merged with Spaces", "Merged with Newlines"]
            plu_file = plu_file.reindex(columns=[*plu_file.columns, *msg_headers])

            # Iterate through PLU file and update the new columns based on MessageNo
            for index, row in plu_file.iterrows():
                message_no = row['MessageNo']
                if message_no > 0 and message_no in id_to_message:
                    message_data = id_to_message[message_no]
                    for i, data in enumerate(message_data):
                        if str(data) != "nan":
                            # Convert the column to 'object' type and update the DataFrame
                            plu_file[msg_headers[i]] = plu_file[msg_headers[i]].astype(object)
                            plu_file.at[index, msg_headers[i]] = str(data)

            return plu_file

        except FileNotFoundError:
            messagebox.showerror("Error", "PLU or Message file not found. Ensure files exist.")
            return None

    try:
        plu_file = pd.read_excel(plu_file_path)
        msg_file = pd.read_excel(msg_file_path)
    except FileNotFoundError:
        messagebox.showerror("Error", "PLU or Message file not found.")
        return

    merged_file = merge_files(plu_file, msg_file)
    if merged_file is None:
        return  # If file merging failed, exit the function

    new_file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")],
        title="Save Merged Excel File"
    )

    if not new_file_path:
        messagebox.showerror("Error", "No save location selected!")
        return

    if os.path.exists(new_file_path):
        overwrite = messagebox.askyesno("Overwrite", f"The file '{new_file_path}' already exists. Do you want to overwrite it?")
        if not overwrite:
            return

    try:
        merged_file.to_excel(new_file_path, index=False)
    except PermissionError:
        messagebox.showerror("Error", f"Permission Error: Close '{new_file_path}' if open.")
        return

    # Applying wrap text formatting in the saved Excel file
    try:
        workbook = openpyxl.load_workbook(new_file_path)
        worksheet = workbook['Sheet1']
        header_row = worksheet[1]
        header_to_column = {cell.value: cell.column for cell in header_row}
        columns_to_wrap = ["Merged", "Merged with Spaces", "Merged with Newlines"]
        for header_name in columns_to_wrap:
            column_index = header_to_column.get(header_name)
            if column_index:
                for cell in worksheet[column_index]:
                    cell.alignment = openpyxl.styles.Alignment(wrapText=True)
        workbook.save(new_file_path)
        messagebox.showinfo("Success", f"Merged and saved to {new_file_path}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while applying formatting: {str(e)}")

def sort_plu_by_gcode():

    plu_file_path = filedialog.askopenfilename(
        title="Select PLU Excel File",
        filetypes=[("Excel Files", "*.xls *.xlsx")]
    )
    
    if not plu_file_path:
        messagebox.showerror("Error", "No PLU file selected!")
        return
    
    plu_name, plu_extension = os.path.splitext(plu_file_path)
    plu_file_name_without_extension = os.path.basename(plu_name)

    if not (plu_extension == ".xlsx" or plu_extension == ".xls"):
        messagebox.showerror("Error", "File type must be an Excel spreadsheet (.xls or .xlsx)")
        return

    try:
        if debug_mode:
            warnings.simplefilter('default')
            warnings.filterwarnings('default')
        else:
            with warnings.catch_warnings():
                warnings.simplefilter(action='ignore', category=DeprecationWarning)
            warnings.filterwarnings("ignore", category=DeprecationWarning)

        new_file_path = filedialog.asksaveasfilename(
            defaultextension=plu_extension,
            filetypes=[("Excel Files", "*.xls *.xlsx")],
            title="Save Sorted Excel File",
            initialfile=plu_file_name_without_extension + " g-sorted"
        )

        if not new_file_path:
            messagebox.showerror("Error", "No save location selected!")
            return

        if os.path.exists(new_file_path):
            overwrite = messagebox.askyesno("Overwrite", f"The file '{new_file_path}' already exists. Do you want to overwrite it?")
            if not overwrite:
                return

        df = pd.read_excel(plu_file_path)
        sorted_df = df.sort_values(by=["GCode", df.columns[0]])

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for r_idx, row in enumerate(dataframe_to_rows(sorted_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        workbook.save(new_file_path)
        messagebox.showinfo("Success", f"Sorted file saved to {new_file_path}")

        if plu_extension == ".xls":
            messagebox.showwarning("Warning", "The older .xls format may cause warnings in Excel. Consider using the .xlsx format for better compatibility.")
    
    except FileNotFoundError:
        messagebox.showerror("Error", f"File '{plu_file_path}' not found or incompatible.")
    except Exception as e:
        messagebox.showerror("Error", f"An unknown error occurred: {str(e)}")

def convert_xls_to_xlsx():
    xls_file_path = filedialog.askopenfilename(
        title="Select .xls File",
        filetypes=[("Old Excel Files", "*.xls")]
    )
    
    if not xls_file_path:
        messagebox.showerror("Error", "No .xls file selected!")
        return

    # Remove .xls extension for creating the new file name
    xls_file_base = xls_file_path.replace(".xls", "").replace(".XLS", "")

    if not os.path.exists(xls_file_path):
        messagebox.showerror("Error", f"File '{xls_file_path}' does not exist or couldn't be found.")
        return

    xlsx_file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")],
        title="Save as .xlsx",
        initialfile=xls_file_base.split('/')[-1] + ".xlsx"
    )
    
    if not xlsx_file_path:
        messagebox.showerror("Error", "No save location selected!")
        return

    try:
        df = pd.read_excel(xls_file_path, sheet_name=None)
        with pd.ExcelWriter(xlsx_file_path, engine='xlsxwriter') as writer:
            for sheet_name, data in df.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)
        messagebox.showinfo("Success", f"Converted and saved to {xlsx_file_path}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

def convert_xlsx_to_xls():
    xlsx_file_path = filedialog.askopenfilename(
        title="Select .xlsx File",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    
    if not xlsx_file_path:
        messagebox.showerror("Error", "No .xlsx file selected!")
        return

    # Remove .xlsx extension for creating the new file name
    xlsx_file_base = xlsx_file_path.replace(".xlsx", "")

    if not os.path.exists(xlsx_file_path):
        messagebox.showerror("Error", f"File '{xlsx_file_path}' does not exist or couldn't be found.")
        return

    xls_file_path = filedialog.asksaveasfilename(
        defaultextension=".xls",
        filetypes=[("Old Excel Files", "*.xls")],
        title="Save as .xls",
        initialfile=xlsx_file_base.split('/')[-1] + ".xls"
    )
    
    if not xls_file_path:
        messagebox.showerror("Error", "No save location selected!")
        return

    try:
        if debug_mode:
            warnings.simplefilter('default')
            warnings.filterwarnings('default')
        else:
            with warnings.catch_warnings():
                warnings.simplefilter(action='ignore', category=DeprecationWarning)
            warnings.filterwarnings("ignore", category=DeprecationWarning)

        workbook_xlsx = openpyxl.load_workbook(xlsx_file_path)
        workbook_xls = xlwt.Workbook()

        for sheet_name in workbook_xlsx.sheetnames:
            sheet_xlsx = workbook_xlsx[sheet_name]
            sheet_xls = workbook_xls.add_sheet(sheet_name)
            for row_idx, row in enumerate(sheet_xlsx.iter_rows(min_row=1, values_only=True), start=1):
                for col_idx, value in enumerate(row, start=1):
                    sheet_xls.write(row_idx - 1, col_idx - 1, value)

        workbook_xls.save(xls_file_path)
        messagebox.showinfo("Success", f"Converted and saved to {xls_file_path}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

def toggle_debug_mode():
    global debug_mode
    debug_mode = not debug_mode
    if debug_mode:
        # Unsuppress FutureWarnings
        warnings.simplefilter('default')
    else:
        # Suppress FutureWarnings
        with warnings.catch_warnings():
            warnings.simplefilter(action='ignore', category=FutureWarning)
    refresh_button_text()

def refresh_button_text():
    debug_button_text = "Set Debug Mode to " + str((not debug_mode))
    debug_button.config(text=debug_button_text)

root = tk.Tk()
root.title("PLU Message File Manager")
root.geometry("600x350")
alt_font = ('Courier', 14, 'bold')

# Create buttons for each task
button_config = [
    ("Convert .msg to Excel", convert_msg_to_excel),
    ("Convert Excel to .msg", convert_excel_to_msg),
    ("Append Message to PLU Excel", append_msg_to_plu),
    ("Sort PLU Excel by g-code", sort_plu_by_gcode),
    ("Convert .xls to .xlsx", convert_xls_to_xlsx),
    ("Convert .xlsx to .xls", convert_xlsx_to_xls)
]

for text, command in button_config:
    tk.Button(root, text=text, command=command, width=40, font=alt_font).pack(pady=5)

# Dynamic button text
debug_button_text = "Set Debug Mode to " + str((not debug_mode))
debug_button = tk.Button(root, text=debug_button_text, command=toggle_debug_mode, width=40, font=alt_font)
debug_button.pack(pady=5)

# Start the Tkinter event loop
root.mainloop()
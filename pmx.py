# romero@engineer.com

# Package Requirements:
# pip install xlsxwriter pandas openpyxl xlwt

# How a message in a .msg file is formatted:
#  ID  M1      M2      M3      M4      M5      M6      M7      M8
# [__][___][_][___][_][___][_][___][_][___][_][___][_][___][_][___][_]
#  4   50   1  50   1  50   1  50   1  50   1  50   1  50   1  50   1

# Convert a .msg file to a spreadsheet file like so:
# - Take in the msg file given by the user
#   - If it exists, begin parsing 412 bytes:
#     - Check the first four bytes to get the message ID number (ID number X)
#       - Number is ABCD, where A is the least significant digit and D is the most
#       - ID = A + (B * 256) + (C * 256^2) + (D * 256^3) where A/B/C/D can range from 0 to 255
#     - Check the next fifty bits for the contents of message line 1 of message number X
#       - Any instance of a NULL byte indicates it is blank from that point on.
#     - Skip NULL byte 1
#     - Repeat for message lines and the individual NULL bytes 2 through 8
#     - Make the message out of the collected data
#     - Add message to list of messages
#       - If there are more bytes, repeat
#       - If there are no more bytes, proceed to export to an Excel spreadhseet
#         - Set header (first row)
#         - For each message
#           - Fill row with ID, M1 through M8, and merged messages
#           - Export as same-named file .xlsx
#         - Check if the IDs are in order
#           - If not in order, export a sorted version
#           - If in order, notify that it already is sorted
#   - If the file does not exist, notify and exit.

# The "problem" characters when using ord():
# 13: Carraige Return     ASCII Code: 10 (Interpretted as Line Feed)
# 128: €                  ASCII Code: 8364
# 129: � (Unused)         ASCII Code: 65533
# 130: ‚                  ASCII Code: 8218
# 131: ƒ                  ASCII Code: 402
# 132: „                  ASCII Code: 8222
# 133: …                  ASCII Code: 8230
# 134: †                  ASCII Code: 8224
# 135: ‡                  ASCII Code: 8225
# 136: ˆ                  ASCII Code: 710
# 137: ‰                  ASCII Code: 8240
# 138: Š                  ASCII Code: 352
# 139: ‹                  ASCII Code: 8249
# 140: Œ                  ASCII Code: 338
# 141: � (Unused)         ASCII Code: 65533
# 142: Ž                  ASCII Code: 381
# 143: � (Unused)         ASCII Code: 65533
# 144: � (Unused)         ASCII Code: 65533
# 145: ‘                  ASCII Code: 8216
# 146: ’                  ASCII Code: 8217
# 147: “                  ASCII Code: 8220
# 148: ”                  ASCII Code: 8221
# 149: •                  ASCII Code: 8226
# 150: –                  ASCII Code: 8211
# 151: —                  ASCII Code: 8212
# 152: ˜                  ASCII Code: 732
# 153: ™                  ASCII Code: 8482
# 154: š                  ASCII Code: 353
# 155: ›                  ASCII Code: 8250
# 156: œ                  ASCII Code: 339
# 157: � (Unused)         ASCII Code: 65533
# 158: ž                  ASCII Code: 382
# 159: Ÿ                  ASCII Code: 376

import xlsxwriter
import os
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import glob
import xlwt
import shlex

encoding = "windows-1252"

message_chunk_bytes = 412
id_bytes = 4
message_line_bytes = 50

def empty_check(x):
    if len(x) > 1:
        return False
    elif len(x) == 1 and ord(x) == 0:
        return True
    else:
        return x == "" or len(x) == 0

def list_characters_with_ascii(input_string):
    lines = []
    for char in input_string:
        ascii_code = ord(char)
        line = f"{char}: {ascii_code}"
        lines.append(line)
    return "\n".join(lines)

class Message:
    def __init__(self, id=0, M1="", M2="", M3="", M4="", M5="", M6="", M7="", M8=""):
        self.id = id
        self.M1 = M1
        self.M2 = M2
        self.M3 = M3
        self.M4 = M4
        self.M5 = M5
        self.M6 = M6
        self.M7 = M7
        self.M8 = M8
        
    def set_id(self, id):
        self.id = id
        
    def set_m1(self, m1):
        self.M1 = m1
        
    def set_m2(self, m2):
        self.M2 = m2
        
    def set_m3(self, m3):
        self.M3 = m3
        
    def set_m4(self, m4):
        self.M4 = m4
        
    def set_m5(self, m5):
        self.M5 = m5
        
    def set_m6(self, m6):
        self.M6 = m6
        
    def set_m7(self, m7):
        self.M7 = m7
        
    def set_m8(self, m8):
        self.M8 = m8

    def valid_message_id(self):
        return self.id > 0
    
    def message_empty(self):
        return self.M1 != "" | self.M2 != "" | self.M3 != "" | self.M4 != "" | self.M5 != "" | self.M6 != "" | self.M7 != "" | self.M8 != ""
    
    def print(self):
        print("Message", self.id, ": M1: \"", self.M1,
              "\"  M2: \"", self.M2,
              "\"  M3: \"", self.M3,
              "\"  M4: \"", self.M4,
              "\"  M5: \"", self.M5,
              "\"  M6: \"", self.M6, 
              "\"  M7: \"", self.M7,
              "\"  M8: \"", self.M8)
    
    def merge_ms(self, separator=""):
        return f"{self.M1}{separator}{self.M2}{separator}{self.M3}{separator}{self.M4}{separator}{self.M5}{separator}{self.M6}{separator}{self.M7}{separator}{self.M8}"#.replace("\x00", "")

    def merge_ms_optimal(self, separator=""):

        merged_message_lines = self.merge_ms(separator)
        
        if separator == "" and len(separator) == 0:
            return merged_message_lines.replace("\x00", "")

        final_empty_flag = False
        
        # Remove any hanging blank lines
        if empty_check(self.M8):
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M8))]
        else:
            final_empty_flag = True
        if empty_check(self.M7) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M7))]
        else:
            final_empty_flag = True
        if empty_check(self.M6) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M6))]
        else:
            final_empty_flag = True
        if empty_check(self.M5) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M5))]
        else:
            final_empty_flag = True
        if empty_check(self.M4) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M4))]
        else:
            final_empty_flag = True 
        if empty_check(self.M3) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M3))]
        else:
            final_empty_flag = True 
        if empty_check(self.M2) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M2))]
        else:
            final_empty_flag = True 
        if empty_check(self.M1) and not final_empty_flag:
            merged_message_lines = merged_message_lines[:-(len(separator) + len(self.M1))]
        else:
            final_empty_flag = True
        
        return merged_message_lines.replace("\x00", "")
            
def get_id_from_bytes(A, B, C, D):
    return ord(A) + (256 * ord(B)) + (256 * 256 * ord(C)) + (256 * 256 * 256 * ord(D))

def get_id_from_ints(A, B, C, D):
    return A + (256 * B) + (256 * 256 * C) + (256 * 256 * 256 * D)

def write_to_file_m_format(file, message):
    if len(message) > message_line_bytes:
        message = message[:message_line_bytes]
    for i in range(message_line_bytes):
        if i < len(message):
            file.write(message[i].encode(encoding))
        else:
            file.write(b'\x00')

# NULL separator to previous message line number guide:
null_separator_pos_prev_m = {
    54: "M1",
    105: "M2",
    156: "M3",
    207: "M4",
    258: "M5",
    309: "M6",
    360: "M7",
    411: "M8"
}

def is_word_in_input(word, text):
    normalized_word = word.lower().replace(" ", "")
    normalized_text = text.lower().replace(" ", "")
    return normalized_word in normalized_text

print("PLU/MSG/Excel Python program started: \n")

debug_mode = False

# original_sf_warning_settings = warnings.simplefilter('default')
# original_fw_warning_settings = warnings.filterwarnings('default')
# original_warning_settings = warnings.filters[:]

while True:

    print("Select a task:")
    print("1.) Convert .msg file to Excel file")
    print("2.) Convert Excel file to .msg file")
    print("3.) Append message Excel file to PLU Excel file")
    print("4.) Sort PLU Excel by g-code")
    print("5.) Convert .xls to .xlsx (Old Excel to New Excel)")
    print("6.) Convert .xlsx to .xls (New Excel to Old Excel)")
    print("7.) Toggle Debug Mode (Currently set to " + str(debug_mode) + ")")
    print("8.) Get or change directory info")
    print("9.) Help")
    print("0.) Exit program\n")

    user_input = str(input("Type a number and press \'enter\' to select an option: "))
    print("")

    if user_input == '1':

        alt_merge_mode = False
        proceed = True
        print("Select a merged message column style:")
        print("1.) Standard string appending: Identical to having typed the merged strings out. (Compatible with the other tools)")
        print("2.) Formula string appending:  Merges strings together using formulas. (Dynamic, but incompatible with the other tools)")
        print("Enter anything else to cancel this action.\n")

        user_mm_input = input("Type a number and press \'enter\' to select an option: ")
        print("")
        if user_mm_input == '1':
            # alt_merge_mode = False
            pass
        elif user_mm_input == '2':
            alt_merge_mode = True
            # print("NOTE: This Excel file will be saved as a .xlsx file. The old .xls file type does not support these formulas.\n")
        else:
            proceed = False
            print("Cancelling task...\n")

        if proceed:
            msg_file_path = str(input("Enter the name of the msg file: "))
            msg_name, msg_extension = os.path.splitext(msg_file_path)
            msg_file_name_without_extension = os.path.basename(msg_name)
            try:
                with open(msg_file_path, 'rb') as file:
                
                    messages = []
                    i = 0 # Byte counter
                    A = B = C = D = 0 # ID
                    message = Message()
                    message_line = ""
                    null_flag = False
                    null_separator = False
                    while True:
                        byte = file.read(1)
                        if not byte:
                            break  # End of file
                        
                        integer_value = ord(byte)
                        byte_char = chr(integer_value)
                        # Position logic
                        if i == message_chunk_bytes: # Finished with message, reset
                            A = B = C = D = i = 0
                            messages.append(message)
                            message = Message()
                            message_line = ""
                            null_flag = False
                        # NULL separator positions: 54, 105, 156, 207, 258, 309, 360, 411
                        if i in null_separator_pos_prev_m:
                            if (not null_flag & integer_value != 0):
                                message_line += byte_char
                            setattr(message, null_separator_pos_prev_m[i], message_line.replace("\x00", ""))
                            null_flag = False
                            message_line = ""
                        if not null_flag:
                            if i == 0:
                                A = integer_value
                            elif i == 1:
                                B = integer_value
                            elif i == 2:
                                C = integer_value
                            elif i == 3:
                                D = integer_value
                                message.set_id(get_id_from_ints(A, B, C, D))
                            elif i not in null_separator_pos_prev_m: 
                                if integer_value == 0:
                                    null_flag = True
                                else:
                                    message_line += byte_char
                        i += 1
            except FileNotFoundError:
              print("Error:", msg_file_path, "not found. Make sure path is correct, file is in the correct directory, or file exists.\n")
              proceed = False

            if proceed:
                if debug_mode:
                    for msgs in messages:
                        Message.print(msgs)
                        print(msgs.id, Message.merge_ms(msgs, "\n"))
                        print(msgs.id, Message.merge_ms_optimal(msgs, "\n"))

                def export_messages_to_excel(messages, file_path):
                
                    workbook = xlsxwriter.Workbook(file_path)
                    worksheet = workbook.add_worksheet()
                    text_wrap_format = workbook.add_format({'text_wrap': True})
                    header = ["ID", "M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "Merged", "Merged with Spaces", "Merged with Newlines"]
                    for col, label in enumerate(header):
                        worksheet.write(0, col, label)

                    if alt_merge_mode:
                        for row, message in enumerate(messages):
                            worksheet.write(row + 1, 0, message.id)
                            for col, value in enumerate([message.M1, message.M2, message.M3, message.M4, message.M5, message.M6, message.M7, message.M8]):
                                worksheet.write(row + 1, col + 1, value, text_wrap_format)
                                # Merged in column J (Alternative methods)
                                formula_j = '=B{row_num} & C{row_num} & D{row_num} & E{row_num} & F{row_num} & G{row_num} & H{row_num} & I{row_num}'.format(row_num=row + 2)
                                #formula_j = '=CONCAT(B{row_num},C{row_num},D{row_num},E{row_num},F{row_num},G{row_num},H{row_num},I{row_num})'.format(row_num=row + 2)
                                worksheet.write_formula(row + 1, 9, formula_j, text_wrap_format)
                                #worksheet.write(row + 1, 9, formula_j, text_wrap_format)
                                #worksheet.write(row + 1, 9, Message.merge_ms_optimal(message), text_wrap_format)

                                # Merged with Spaces in column K (Alternative methods)
                                formula_k = '=B{row_num} & " " & C{row_num} & " " & D{row_num} & " " & E{row_num} & " " & F{row_num} & " " & G{row_num} & " " & H{row_num} & " " & I{row_num}'.format(row_num=row + 2)
                                #formula_k = '=CONCAT(B{row_num}," ",C{row_num}," ",D{row_num}," ",E{row_num}," ",F{row_num}," ",G{row_num}," ",H{row_num}," ",I{row_num})'.format(row_num=row + 2)
                                worksheet.write_formula(row + 1, 10, formula_k, text_wrap_format)
                                #worksheet.write(row + 1, 10, formula_k, text_wrap_format)
                                #worksheet.write(row + 1, 10, Message.merge_ms_optimal(message, " "), text_wrap_format)

                                # Merged with Newlines in column L (Alternative methods)
                                formula_l = '=B{row_num} & CHAR(10) & C{row_num} & CHAR(10) & D{row_num} & CHAR(10) & E{row_num} & CHAR(10) & F{row_num} & CHAR(10) & G{row_num} & CHAR(10) & H{row_num} & CHAR(10) & I{row_num}'.format(row_num=row + 2)
                                #formula_l = '=CONCAT(B{row_num},CHAR(10),C{row_num},CHAR(10),D{row_num},CHAR(10),E{row_num},CHAR(10),F{row_num},CHAR(10),G{row_num},CHAR(10),H{row_num},CHAR(10),I{row_num})'.format(row_num=row + 2)
                                worksheet.write_formula(row + 1, 11, formula_l, text_wrap_format)
                                #worksheet.write(row + 1, 11, formula_l, text_wrap_format)
                                #worksheet.write(row + 1, 11, Message.merge_ms_optimal(message, "\n"), text_wrap_format)
                    else:
                        for row, message in enumerate(messages):
                            worksheet.write(row + 1, 0, message.id)
                            for col, value in enumerate([message.M1, message.M2, message.M3, message.M4, message.M5, message.M6, message.M7, message.M8, Message.merge_ms_optimal(message, ""), Message.merge_ms_optimal(message, " "), Message.merge_ms_optimal(message, "\n")]):
                                if debug_mode:
                                    print("Merged:", Message.merge_ms_optimal(message, ""))
                                    print("Merged w/ space:", Message.merge_ms_optimal(message, " "))
                                    #print("Merged w/ newline:", Message.merge_ms_optimal(message, "\n"))
                                worksheet.write(row + 1, col + 1, value, text_wrap_format)
                    workbook.close()
                
                if os.path.exists(msg_file_name_without_extension + ".xlsx"):
                    user_ow_input = input(f"The file \'{msg_file_name_without_extension + ".xlsx"}\' already exists. Do you want to overwrite it? (y/n): ").lower()
                    if user_ow_input == 'y' or user_ow_input == "yes":
                        try:
                            os.remove(msg_file_name_without_extension + ".xlsx")
                            print(f"Overwriting {msg_file_name_without_extension + ".xlsx"} ...")
                        except FileNotFoundError:
                            print(f"Error: File {msg_file_name_without_extension + ".xlsx"} not found.")
                        except PermissionError:
                            print(f"Error: Permission denied to delete file {msg_file_name_without_extension + ".xlsx"}.")
                        except Exception as e:
                            print(f"Error: An unknown error occurred. Exception: {e}")
                    elif user_ow_input == 'n' or user_ow_input == "no":
                        print("Overwriting canceled.")
                        proceed = False
                    else:
                        print("Unknown input. Overwriting canceled. The existing file was not overwritten.")
                        proceed = False
                else:
                    print("Exporting to", (msg_file_name_without_extension + ".xlsx"))
                    export_messages_to_excel(messages, (msg_file_name_without_extension + ".xlsx"))

                if proceed:
                    is_sorted = True
                    prev_id = 0
                    for msg_ids in messages:
                        if (msg_ids.id > prev_id):
                            prev_id = msg_ids.id
                        else:
                            is_sorted = False
                            break    
                    if (is_sorted):
                        print("(This file is already sorted by ID\n)")
                    else:
                        sorted_messages = sorted(messages, key=lambda message: message.id)
                        print("Exporting to sorted version to", (msg_file_name_without_extension + " sorted.xlsx\n"))
                        export_messages_to_excel(sorted_messages, (msg_file_name_without_extension + " sorted.xlsx") )
                    print("NOTE: To properly display the merged-text results, select the columns and use the \"Wrap Text\" option.")
                    print("      You may need to change the vertical Align settings (use top or middle) as well if using formulas.\n")

    if user_input == '2':
        print("NOTE: This program is set up to take Excel spreadsheets formatted like so:")
        print("  A      B      C      D      E      F      G      H      I")
        print("[ ID ] [ M1 ] [ M2 ] [ M3 ] [ M4 ] [ M5 ] [ M6 ] [ M7 ] [ M8 ]\n")

        proceed = True

        if debug_mode:
            # Unsuppress DeprecationWarning
            warnings.simplefilter('default')
            warnings.filterwarnings('default')
        else:
            # Suppress DeprecationWarning
            with warnings.catch_warnings():
                warnings.simplefilter(action='ignore', category=DeprecationWarning)
            warnings.filterwarnings("ignore", category=DeprecationWarning)

        excel_file_path = str(input("Enter the name of the Excel file: "))
        excel_name, excel_extension = os.path.splitext(excel_file_path)
        excel_file_name_without_extension = os.path.basename(excel_name)

        new_file_path = excel_file_name_without_extension + ".msg"
        if os.path.exists(new_file_path):
            user_ow_input = input(f"The file \'{new_file_path}\' already exists. Do you want to overwrite it? (y/n): ").lower()
            if user_ow_input == 'y' or user_ow_input == "yes":
                try:
                    os.remove(new_file_path)
                    print(f"Overwriting {new_file_path} ...")
                except FileNotFoundError:
                    print(f"Error: File {new_file_path} not found.")
                except PermissionError:
                    print(f"Error: Permission denied to delete file {new_file_path}.")
                    proceed = False
                except Exception as e:
                    print(f"Error: An unknown error occurred. Exception: {e}")
                    proceed = False
            elif user_ow_input == 'n' or user_ow_input == "no":
                print("Overwriting canceled.")
                proceed = False
            else:
                print("Unknown input. Operation and overwriting canceled. The existing file was not overwritten.")
                proceed = False
        
        if proceed:
            try:
                df = pd.read_excel(excel_file_path)

                # Drop rows with missing ID (this would cause issues for the .msg file)
                df = df.dropna(subset=['ID'])
                rows = 0
                messages = []
                i = 0 # Byte counter
                e_message = Message()
                message_line = ""
                ids = []

                for index, row in df.iterrows():
                    e_message_id = int(row['ID'])
                    e_ms = [str(row[f'M{i}']) if pd.notna(row[f'M{i}']) else '' for i in range(1, 9)]
                    if len(e_ms) > message_line_bytes: # Message lines cannot exceed 50 characters
                        e_ms = e_ms[:message_line_bytes]
                    ids.append(e_message_id)
                    messages.append(Message(e_message_id, *e_ms))
                    if debug_mode:
                        print(f"[{rows}] Message ID: {e_message_id}")
                        print(f"[{rows}] Messages 1-8: {e_ms}")

                used_ids = []
                repeat_ids = []

                if len(messages) > 0:
                    with open(new_file_path, 'wb') as file:
                        for msg in messages:
                            A = B = C = D = 0 # ID
                            if debug_mode:
                                print(f"Message Info: [{msg.id}] \"{msg.M1}\" \"{msg.M2}\" \"{msg.M3}\" \"{msg.M4}\" \"{msg.M5}\" \"{msg.M6}\" \"{msg.M7}\" \"{msg.M8}\"")
                                print(f"Lengths: M1:{len(msg.M1)}  M2:{len(msg.M2)}  M3:{len(msg.M3)}  M4:{len(msg.M4)}  M5:{len(msg.M5)}  M6:{len(msg.M6)}  M7:{len(msg.M7)}  M8:{len(msg.M8)}")
                            if(msg.id not in used_ids):
                                if msg.id > 4294967295:
                                    msg.id = 4294967295
                                elif msg.id < 1:
                                    msg.id = 1
                                D = msg.id // (256 ** 3)
                                C = (msg.id % (256 ** 3)) // (256 ** 2)
                                B = (msg.id % (256 ** 2)) // 256
                                A = msg.id % 256
                                if debug_mode:
                                    print(f"Message ID Base-10 to Base-256: {D} | {C} | {B} | {A}\n")
                                #412: 4 50 1 50 1 50 1 50 1 50 1 50 1 50 1 50 1
                                file.write(bytes([A]))                     # 1    |     1
                                file.write(bytes([B]))                     # 1    |     2
                                file.write(bytes([C]))                     # 1    |     3
                                file.write(bytes([D]))                     # 1    |     4
                                write_to_file_m_format(file, msg.M1)       # 50   |    54
                                file.write(b'\x00')                        # 1    |    55
                                write_to_file_m_format(file, msg.M2)       # 50   |   105
                                file.write(b'\x00')                        # 1    |   106
                                write_to_file_m_format(file, msg.M3)       # 50   |   156
                                file.write(b'\x00')                        # 1    |   157
                                write_to_file_m_format(file, msg.M4)       # 50   |   207
                                file.write(b'\x00')                        # 1    |   208
                                write_to_file_m_format(file, msg.M5)       # 50   |   258
                                file.write(b'\x00')                        # 1    |   259
                                write_to_file_m_format(file, msg.M6)       # 50   |   309
                                file.write(b'\x00')                        # 1    |   310
                                write_to_file_m_format(file, msg.M7)       # 50   |   360
                                file.write(b'\x00')                        # 1    |   361
                                write_to_file_m_format(file, msg.M8)       # 50   |   411
                                file.write(b'\x00')                        # 1    |   412

                                used_ids.append(msg.id)
                            else:
                                repeat_ids.append(msg.id)

                if (debug_mode):
                    print(f"Total: {len(ids)} Used: {len(used_ids)} Repeat: {len(repeat_ids)}\n")                
                if len(used_ids) < len(ids) or repeat_ids > 0:
                    if len(ids) - len(used_ids) == 1 or repeat_ids == 1:
                        print(f"!!! ALERT: 1 message had a reused ID and is not included in this PLU message file.")
                    else:
                        print(f"!!! ALERT: {len(repeat_ids)} messages had reused IDs and are not included in this PLU message file.")
                    print("*** Verify that each message has a unique non-zero ID in the spreadsheet. This may just be a duplicate entry glitch.")
                    print("Repeat ID(s):", repeat_ids, "\n")

                print("Saved to " + excel_file_name_without_extension + ".msg\n")
            except FileNotFoundError:
                  print("Error:", excel_file_path, "not found. Make sure path is correct, file is in the correct directory, or file exists.\n")
                  proceed = False

    if user_input == '3':
        proceed = True
        if debug_mode:
            # Unsuppress FutureWarnings
            warnings.simplefilter('default')
        else:
            # Suppress FutureWarnings
            with warnings.catch_warnings():
                warnings.simplefilter(action='ignore', category=FutureWarning)
        
        print("NOTE: This merging action does NOT support merging formulas.")
        print("Merged-message cells using formulas will say \'0\' instead, but all message lines (M1 - M8) will remain intact.\n")

        plu_file_path = str(input("Enter the PLU (.xlsx) file to add the messages to: "))
        plu_name, plu_extension = os.path.splitext(plu_file_path)
        plu_file_name_without_extension = os.path.basename(plu_name)
        print("")
        msg_file_path = str(input("Enter the msg (.xlsx) file be added: "))
        msg_name, msg_extension = os.path.splitext(msg_file_path)
        msg_file_name_without_extension = os.path.basename(msg_name)
        print("")

        def merge_files(plu_file, msg_file):
            try:
                # Create a dictionary to map IDs to corresponding message data
                id_to_message = {row['ID']: row[1:] for _, row in msg_file.iterrows()}

                msg_headers = ["M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "Merged", "Merged with Spaces", "Merged with Newlines"]
                plu_file = plu_file.reindex(columns=[*plu_file.columns, *msg_headers])

                # Iterate through plu_file and update the new columns based on MessageNo
                for index, row in plu_file.iterrows():
                    message_no = row['MessageNo']
                    if message_no > 0 and message_no in id_to_message:
                        message_data = id_to_message[message_no]
                        for i, data in enumerate(message_data):
                            if str(data) != "nan":
                                #plu_file.at[index, msg_headers[i]] = str(data)
                                # Convert the column to 'object' type
                                plu_file[msg_headers[i]] = plu_file[msg_headers[i]].astype(object)
                                # Update the DataFrame
                                plu_file.at[index, msg_headers[i]] = str(data)

                # In an older version, for some reason the regular "Merged" column doesn't work without doing this
                #plu_file['Merged'] = plu_file.apply(lambda row: ''.join(str(value) if not pd.isna(value) else '' for value in row['M1':'M8']), axis=1)

                # Enabling Text wrapping for the merged lines
                #for column in ['R', 'S', 'T']:
                #    for cell in worksheet[column]:
                #        cell.alignment = openpyxl.styles.Alignment(wrapText=True)

                return plu_file

            except FileNotFoundError:
                print("Error:", plu_file_path, "or", msg_file_path, "not found. Make sure path is correct, file is in the correct directory, or file exists.")
                proceed = False

        if proceed:
            try:
                # Load the first Excel file (plu_file_path)
                plu_file = pd.read_excel(plu_file_path)

                # Load the second Excel file (msg_file_path)
                msg_file = pd.read_excel(msg_file_path)

            except FileNotFoundError:
                print("Error:", plu_file_path, "or", msg_file_path, "not found. Make sure path is correct, file is in the correct directory, or file exists.")
                proceed = False

        if proceed:
            merged_file = merge_files(plu_file, msg_file)

            new_file_path = plu_file_name_without_extension + '+' + msg_file_name_without_extension + '.xlsx'

            if os.path.exists(new_file_path):
                user_ow_input = input(f"The file \'{new_file_path}\' already exists. Do you want to overwrite it? (y/n): ").lower()
                if user_ow_input == 'y' or user_ow_input == "yes":
                    try:
                        os.remove(new_file_path)
                        print(f"Overwriting {new_file_path} ...")
                    except FileNotFoundError:
                        print(f"Error: File {new_file_path} not found.")
                    except PermissionError:
                        print(f"Error: Permission denied to delete file {new_file_path}.")
                        proceed = False
                    except Exception as e:
                        print(f"Error: An unknown error occurred. Exception: {e}")
                        proceed = False
                elif user_ow_input == 'n' or user_ow_input == "no":
                    print("Overwriting canceled.")
                    proceed = False
                else:
                    print("Unknown input. Operation and overwriting canceled. The existing file was not overwritten.")
                    proceed = False

        if proceed:    
            try:
                merged_file.to_excel(new_file_path, index=False)
            except PermissionError:
                print("Permission Error: Close", new_file_path, "if open.")
                proceed = False

        if proceed:
            workbook = openpyxl.load_workbook(new_file_path)
            worksheet = workbook['Sheet1']
            header_row = worksheet[1]
            header_to_column = {cell.value: cell.column for cell in header_row}
            columns_to_wrap = ["Merged", "Merged with Spaces", "Merged with Newlines"]
            for header_name in columns_to_wrap:
                column_index = header_to_column.get(header_name)
                if column_index:
                    for cell in worksheet[column_index]:
                        cell.alignment = openpyxl.styles.Alignment(wrapText=True)
            workbook.save(new_file_path)
            print(f"Merged and saved to {new_file_path}.\n")


    if user_input == '4':
        proceed = True
        plu_file_path = str(input("Enter the PLU file (.xls or .xlsx) to be GCode-sorted: "))
        print("")
        if plu_file_path.endswith(".xlsx") or plu_file_path.endswith(".xls"):
            plu_name, plu_extension = os.path.splitext(plu_file_path)
            plu_file_name_without_extension = os.path.basename(plu_name)
        else:
            print("Error: File type must be an Excel spreadsheet. (.xls or .xlsx)")
        try:
            if debug_mode:
                # Unsuppress DeprecationWarning
                # warnings.simplefilter(**original_sf_warning_settings)
                # warnings.filterwarnings(**original_fw_warning_settings)
                warnings.simplefilter('default')
                warnings.filterwarnings('default')
            else:
                # Suppress DeprecationWarning
                with warnings.catch_warnings():
                    warnings.simplefilter(action='ignore', category=DeprecationWarning)
                warnings.filterwarnings("ignore", category=DeprecationWarning)

            new_file_path = plu_file_name_without_extension + " g-sorted" + plu_extension
            if os.path.exists(new_file_path):
                user_ow_input = input(f"The file \'{new_file_path}\' already exists. Do you want to overwrite it? (y/n): ").lower()
                if user_ow_input == 'y' or user_ow_input == "yes":
                    try:
                        os.remove(new_file_path)
                        print(f"Overwriting {new_file_path} ...")
                    except FileNotFoundError:
                        print(f"Error: File {new_file_path} not found.")
                    except PermissionError:
                        print(f"Error: Permission denied to delete file {new_file_path}.")
                        proceed = False
                    except Exception as e:
                        print(f"Error: An unknown error occurred. Exception: {e}")
                        proceed = False
                elif user_ow_input == 'n' or user_ow_input == "no":
                    print("Overwriting canceled.")
                    proceed = False
                else:
                    print("Unknown input. Operation and overwriting canceled. The existing file was not overwritten.")
                    proceed = False
            
            if proceed:
                df = pd.read_excel(plu_file_path)
                sorted_df = df.sort_values(by=["GCode", df.columns[0]])
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                for r_idx, row in enumerate(dataframe_to_rows(sorted_df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                print("Sorted successfully. Saving to " + (plu_file_name_without_extension + " g-sorted" + plu_extension) + "\n")
                
                workbook.save(plu_file_name_without_extension + " g-sorted" + plu_extension)
                if plu_extension == ".xls":
                    print("NOTE: This program works better with the newer .xlsx Excel file format. The older .xls format can work with this, but the created file may give a warning about data being corrupted or unsafe.\n")
        except FileNotFoundError:
            print("Error:", (plu_file_path), "not found or incompatible. Make sure path is correct, file type is correct, file is in the correct directory, or file exists.\n")

    if user_input == '5': # OLD -> NEW
        print("NOTE: This does NOT delete or overwrite the original Excel file. This command makes a copy of the file with a different file extension.\n")
        xls_file_path = input("Enter the name of the .xls file to be converted: ")
        print("")
        xls_file_path = xls_file_path.replace(".xls", "")
        if os.path.exists(xls_file_path + ".xls"):
            xlsx_file_path = xls_file_path + '.xlsx'
            df = pd.read_excel(xls_file_path + ".xls", sheet_name=None)
            with pd.ExcelWriter(xlsx_file_path, engine='xlsxwriter') as writer:
                for sheet_name, data in df.items():
                    data.to_excel(writer, sheet_name=sheet_name, index=False)

            print("Converted and saved to "+ xlsx_file_path+ "\n")
        else:
            print("Error: File \'" + xls_file_path + ".xls\' does not exist or couldn't be found.\n")

    if user_input == '6': # NEW -> OLD
        print("NOTE: This does NOT delete or overwrite the original Excel file. This command makes a copy of the file with a different file extension.")
        print("ALSO NOTE: As .xls is an older format, many formulas do NOT work when converted from .xlsx to .xls.\n")
        xlsx_file_path = input("Enter the name of the .xlsx file: ")
        print("")
        xlsx_file_path = xlsx_file_path.replace(".xlsx", "")
        if os.path.exists(xlsx_file_path + ".xlsx"):
            if debug_mode:
                # Unsuppress DeprecationWarning
                warnings.simplefilter('default')
                warnings.filterwarnings('default')
            else:
                # Suppress DeprecationWarning
                with warnings.catch_warnings():
                    warnings.simplefilter(action='ignore', category=DeprecationWarning)
                warnings.filterwarnings("ignore", category=DeprecationWarning)

            xls_file_path = xlsx_file_path + '.xls'
            workbook_xlsx = openpyxl.load_workbook(xlsx_file_path + ".xlsx")
            workbook_xls = xlwt.Workbook()
            for sheet_name in workbook_xlsx.sheetnames:
                sheet_xlsx = workbook_xlsx[sheet_name]
                sheet_xls = workbook_xls.add_sheet(sheet_name)
                for row_idx, row in enumerate(sheet_xlsx.iter_rows(min_row=1, values_only=True), start=1):
                    for col_idx, value in enumerate(row, start=1):
                        sheet_xls.write(row_idx - 1, col_idx - 1, value)
            workbook_xls.save(xls_file_path)
            print("Converted and saved to "+ xls_file_path + "\n")
        else:
            print("Error: File \'" + xlsx_file_path + ".xlsx\' does not exist or couldn't be found.\n")

    if user_input == '7' or is_word_in_input('debug', user_input):
        debug_mode = not debug_mode
        if debug_mode:
            print("Debug Mode enabled\n")
        else:
            print("Debug Mode disabled\n")

    elif user_input == '8':
        print("Your current working directory is: " + str(os.getcwd()) + "\n")
        
        user_nd_input = input("Enter a new directory and press \'enter\' if you wish to change it (leave blank to stay): ")
        print("")
        if user_nd_input != "" or len(user_nd_input) != 0:
            try:
                os.chdir(user_nd_input)
                print("Your current working directory is now: " + str(os.getcwd())+ "\n")
            except FileNotFoundError:
                print(f"Error: Directory \'{user_nd_input}\' was not found.\n")
                print("Your current working directory is: " + str(os.getcwd()) + "\n")
            except Exception as e:
                print(f"Error: An unknown error has occurred.\n Input: \'\'{user_nd_input}\'\'\n Exception: {e}\n")
                print("Your current working directory is: " + str(os.getcwd()) + "\n")

        msg_files = glob.glob("*.msg")
        if len(msg_files) == 0:
            print("There are no PLU message files in the current working directory.")
        elif len(msg_files) == 1:
            print("There is 1 PLU message file in the current working directory:\n")
            print(msg_files[0])
        else:
            print("There are " + str(len(msg_files)) + " PLU message files in the current working directory:\n")
            for file in msg_files:
                print(file)
        
        print("")

        xlsx_files = glob.glob("*.xlsx")
        xls_files = glob.glob("*.xls")

        if (len(xlsx_files) + len(xls_files)) == 0:
            print("There are no Excel files in the current working directory.\n")
        elif (len(xlsx_files) + len(xls_files)) == 1:
            print("There is 1 Excel file in the current working directory:\n")
            if (len(xlsx_files) > len(xls_files)):
                print(xlsx_files[0])
            else:
                print(xls_files[0])
        else:
            print("There are " + str((len(xlsx_files) + len(xls_files))) + " Excel files in the current working directory:\n")
            if (len(xlsx_files) > 0):
                for file in xlsx_files:
                    print(file)
                print("")
            if (len(xls_files) > 0):
                for file in xls_files:
                    print(file)
                print("")

    elif user_input == '9' or is_word_in_input('help', user_input):
        print("`7MM\"\"\"Mq.`7MMM.     ,MMF'`YMM'   `MP\'")        
        print("  MM   `MM. MMMb    dPMM    VMb.  ,P")          
        print("  MM   ,M9  M YM   ,M MM     `MM.M\'        gp") 
        print("  MMmmdM9   M  Mb  M\' MM       MMb         \"\"") 
        print("  MM        M  YM.P\'  MM     ,M'`Mb.") 
        print("  MM        M  `YM\'   MM    ,P   `MM.      ,,") 
        print(".JMML.    .JML. `\'  .JMML..MM:.  .:MMa.    db      PLU-Message-Excel File Manager for LP-Works\n")

        print("This is the PLU-msg-Excel file converter and merger version 1.0.0\n")
        print("This program is meant to facilitate the usage of the LP-Works program for CAS scales by providing the following utilities:\n")
        
        print("1.) A PLU message file to Excel file converter. This takes in a PLU message file (.msg) and converts it to an Excel spreadsheet (.xlsx).")
        print("Each message in a message file is broken up into eight separate messages of 50 characters, so the spreadsheet shows each message line and the merged results (no space between, a space between, and a newline between) of each message for each corresponding message ID.\n")

        print("2.) An Excel file to PLU message file converter. This takes in an Excel spreadsheet (.xlsx) and converts it to a PLU message file (.msg).")
        print("THIS REQUIRES SPECIFIC FORMATTING: [ID M1 M2 ... M8] For each ID in the first column, the first 50 character contents for each of the eight message lines will be merged to form the contents of the full message.\n")

        print("3.) A message file spreadsheet to PLU file spreadsheet appender. This takes the produced message file spreadsheet and appends it to a PLU spreadsheet by matching the message ID to the MessageNo column.")
        print("Each PLU has a MessageNo, which can be 0 (no message) or some number X. If there is an X-ID message, it will add that message info (M1 through M8 and the merged results) to the columns after PLU information.\n")

        print("4.) A PLU spreadsheet g-code sorter. This takes a PLU spreadsheet and sorts the PLU entries by its g-code.")
        print("Each PLU entry has a g-code. If modifications need to be made on certain g-coded entries, this can facilitate the search for those items.\n")

        print("5.) A .xls to .xlsx converter. Takes a pre-2007 format Excel spreadsheet file and converts it and its contents to the newer 2007 format Excel spreadsheet file.\n")

        print("6.) A .xlsx to .xls converter. Takes the newer 2007 format Excel spreadsheet file and converts it and its contents to the older pre-2007 Excel spreadsheet file, compatible with LP-Works, though some .xlsx files can work with LP-Works.\n")

        print("There is also a 7.) \"debug mode\" option for testing purposes and a 8.) \"get / change directory info\" command to see and set the current working directory and view compatible files.\n")

    elif user_input == '0' or is_word_in_input('exit', user_input) or is_word_in_input('quit', user_input):
        # print("Exiting program.")
        break  # Exit while True loop

    elif user_input.lower().startswith("cd ") and len(user_input) > 3:
        os.chdir(user_input[3:])
    elif user_input.lower() == "pwd":
        print(str(os.getcwd()), "\n")
    elif is_word_in_input('ls', user_input) or is_word_in_input('dir', user_input):
        files = os.listdir()
        for file in files:
            print(file)
    elif user_input.lower().startswith("mkdir ") and len(user_input) > 6:
        newdir = user_input[6:]
        try:
            os.makedirs(newdir)
            print(f"Directory '{newdir}' created successfully.\n")
        except FileExistsError:
            print(f"Error: Directory '{newdir}' already exists.\n")
        except Exception as e:
            print(f"Error: An unknown error occurred. Exception: {e}\n")
    elif user_input.lower().startswith("mv ") and len(user_input) > 5:
        mv_inputs = shlex.split(user_input[3:])
        if len(mv_inputs) == 2:
            old_filename, new_filename = mv_inputs
            try:
                os.rename(old_filename, new_filename)
                print(f"File \'{old_filename}\' renamed to \'{new_filename}\' successfully.\n")
            except FileNotFoundError:
                print(f"Error: File \'{old_filename}\' not found.\n")
            except PermissionError:
                print(f"Error: Permission denied to rename file \'{old_filename}\' or file is currently open.\n")
            except Exception as e:
                print(f"Error: An unknown error occurred. Exception: {e}\n")
        else:
            print(f"Error: 2 inputs required: The old filename and a new filename. Inputs: {mv_inputs}\n")

    else:
        if not (user_input.isdigit() and len(user_input) == 1):
            print(f"Unrecognized input: {user_input}\n")

    input("Press \'enter\' to continue...")
    print("")

# Out of while True loop:
print("PLU/MSG/Excel Python program exited.\n")